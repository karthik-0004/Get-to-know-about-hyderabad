"""
Housing.com - Hyderabad APARTMENT Scraper (via Property Type Filter)
=====================================================================
FIXES v3:
  1. is_valid_apartment_url() uses path-segment comparison — the apartment
     URL's last path segment must match the locality URL's last segment.
     This catches ALL bad redirects, not just /hyderabad/hyderabad.
  2. MAX_PAGES_ABSOLUTE = 20: hard ceiling per locality, fires before
     every page load so it can never be bypassed.
  3. Sanity check: if reported total > 500 but page-1 cards < 30, override.
  4. Dynamic cap = min(ceil(total/30)+1, MAX_PAGES_ABSOLUTE).

Output: apartment_filter.xlsx

Setup:
    pip install playwright openpyxl
    playwright install chromium

Run:
    python apartment_filter_scraper.py
"""

import re
import time
import os
import math
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ───────────────────────────────────────────────────────────────────
BASE_URL           = "https://housing.com/price-trends/property-rates-for-buy-in-hyderabad_telangana-P679xe73u28050522"
TOTAL_PAGES        = 17
OUTPUT_FILE        = "apartment_filter.xlsx"
HEADLESS           = False
MAX_PAGES_ABSOLUTE = 20      # hard ceiling per locality — no matter what

# ── EXCEL ─────────────────────────────────────────────────────────────────────
COLUMNS    = ["Locality", "Avg Price/Sqft", "Price Min", "Price Max",
              "BHK", "Sqft", "Price", "Amenities"]
COL_WIDTHS = [25, 18, 16, 16, 10, 12, 15, 55]
HDR_FILL   = PatternFill("solid", fgColor="006064")
HDR_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ALT_FILL   = PatternFill("solid", fgColor="E0F7FA")
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── URL VALIDATION ────────────────────────────────────────────────────────────

def last_path_segment(url: str) -> str:
    """Return the final non-empty path segment of a URL, lowercased."""
    path = urlparse(url.rstrip("/")).path
    parts = [p for p in path.split("/") if p]
    return parts[-1].lower() if parts else ""


def is_valid_apartment_url(apt_url: str, locality_url: str) -> bool:
    """
    Return True only if apt_url is a real locality-level apartment URL.

    Rules:
      - Must not be the bare city root.
      - The last path segment of apt_url must match the last segment of
        locality_url (allows suffix like '-apartment' or '-flat').
        e.g.  locality ends in 'maripally'       → apt must end in 'maripally'
              locality ends in 'asif_nagar_north' → apt must end in 'asif_nagar_north'
    """
    apt = apt_url.rstrip("/").lower()

    bad_exact = {
        "https://housing.com/in/buy/hyderabad/hyderabad",
        "https://housing.com/in/buy/hyderabad",
    }
    if apt in bad_exact:
        return False

    loc_seg = last_path_segment(locality_url)
    apt_seg = last_path_segment(apt_url)

    if not loc_seg or not apt_seg:
        return False

    # Exact match or one is a prefix of the other (e.g. slug + '-apartment')
    if loc_seg == apt_seg:
        return True
    if apt_seg.startswith(loc_seg) or loc_seg.startswith(apt_seg):
        return True

    return False   # segments differ → bad redirect


# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────

def init_wb():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        print(f"↩  Resuming {OUTPUT_FILE}  ({ws.max_row - 1} rows saved)")
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apartment"
    for ci, (h, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"✔  Created {OUTPUT_FILE}")
    return wb, ws


def save_locality(wb, ws, locality, avg_price, p_min, p_max, props):
    if not props:
        props = [{"bhk": "", "sqft": "", "price": "", "amenities": ""}]
    for prop in props:
        r = ws.max_row + 1
        vals = [locality, avg_price, p_min, p_max,
                prop["bhk"], prop["sqft"], prop["price"], prop["amenities"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.border = BORDER
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                c.fill = ALT_FILL
    wb.save(OUTPUT_FILE)
    print(f"   💾 Saved {len(props)} rows for '{locality}'  (total: {ws.max_row - 1})")


def get_saved_localities(ws):
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    return done


def safe_text(el):
    try:
        return el.inner_text().strip() if el else ""
    except:
        return ""


def scroll_full(page):
    for pct in [0.3, 0.6, 1.0]:
        page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pct})")
        time.sleep(0.8)


def load_page(page, url, retries=3):
    for attempt in range(retries):
        try:
            page.goto(url, timeout=50000, wait_until="domcontentloaded")
            time.sleep(4)
            return True
        except PWTimeout:
            print(f"   Timeout attempt {attempt+1}/{retries}, retrying...")
            time.sleep(3)
    print(f"   ✗ Failed: {url}")
    return False


# ── GET APARTMENT URL FROM PROPERTY TYPE DROPDOWN ────────────────────────────

def get_apartment_url(page, locality_url):
    print(f"      🔍 Finding Apartment URL: {locality_url[:70]}")

    if not load_page(page, locality_url):
        return locality_url

    clicked = page.evaluate("""
        () => {
            const els = document.querySelectorAll('div.css-r9ub9f');
            for (const el of els) {
                if (el.innerText.trim() === 'Property Type') {
                    el.click(); return true;
                }
            }
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while (node = walker.nextNode()) {
                if (node.textContent.trim() === 'Property Type') {
                    node.parentElement.click(); return true;
                }
            }
            return false;
        }
    """)

    if not clicked:
        print(f"      ⚠ Could not open Property Type dropdown — using locality URL")
        return locality_url

    print(f"      ✔ Opened Property Type dropdown")
    time.sleep(1.5)

    apt_href = page.evaluate("""
        () => {
            const options = document.querySelectorAll('div.option, div[class*="option"]');
            for (const opt of options) {
                const txt = opt.innerText.trim();
                if (txt === 'Apartment' || txt === 'Apartments' ||
                    txt === 'Flat'      || txt === 'Flats') {
                    const link = opt.querySelector('a');
                    if (link) return link.getAttribute('href');
                }
            }
            for (const a of document.querySelectorAll('a')) {
                const txt = a.innerText.trim();
                if (txt === 'Apartment' || txt === 'Apartments' ||
                    txt === 'Flat'      || txt === 'Flats') {
                    return a.getAttribute('href');
                }
            }
            return null;
        }
    """)

    if apt_href:
        full_url = ("https://housing.com" + apt_href
                    if apt_href.startswith("/") else apt_href)

        if is_valid_apartment_url(full_url, locality_url):
            print(f"      ✔ Apartment URL: {full_url}")
            return full_url
        else:
            print(f"      ⚠ Dropdown returned mismatched URL:")
            print(f"         locality : .../{last_path_segment(locality_url)}")
            print(f"         got      : .../{last_path_segment(full_url)}")
            print(f"         → Falling back to locality URL")
            page.keyboard.press("Escape")
            return locality_url

    print(f"      ⚠ Apartment link not found — using locality URL")
    page.keyboard.press("Escape")
    return locality_url


# ── TABLE: locality rows ──────────────────────────────────────────────────────

def get_locality_rows(page):
    rows_data = []
    rows = page.query_selector_all("div.css-1s17y02")
    if not rows:
        rows = [r for r in page.query_selector_all("div[class^='css-']")
                if r.query_selector("a[href*='property-rates']")]
    print(f"   Found {len(rows)} locality rows")

    for row in rows:
        try:
            name_el  = (row.query_selector("a.css-673lf3") or
                        row.query_selector("a[href*='property-rates']") or
                        row.query_selector("a"))
            locality = safe_text(name_el)
            if not locality:
                continue

            price_el    = (row.query_selector("span.css-69n8oe") or
                           row.query_selector("span:nth-child(2)"))
            range_el    = (row.query_selector("span.css-5sq9yq") or
                           row.query_selector("span:nth-child(3)"))
            avg_price   = safe_text(price_el)
            price_range = safe_text(range_el)
            parts = re.split(r"\s*[-–]\s*", price_range)
            p_min = parts[0].strip() if parts else ""
            p_max = parts[1].strip() if len(parts) > 1 else ""

            link_el  = (row.query_selector("span.css-15j6032 a") or
                        row.query_selector("a[href*='/in/buy/']"))
            href     = link_el.get_attribute("href") if link_el else ""
            prop_url = ("https://housing.com" + href if href.startswith("/")
                        else href if href.startswith("http") else "")

            rows_data.append({
                "locality":  locality,
                "avg_price": avg_price,
                "p_min":     p_min,
                "p_max":     p_max,
                "prop_url":  prop_url,
            })
        except Exception as e:
            print(f"   Row parse error: {e}")
    return rows_data


# ── PARSE CARDS ───────────────────────────────────────────────────────────────

def extract_fields(txt):
    if not txt or len(txt) < 15:
        return None

    bhk_m = re.search(r"([\d.,\s&]+\s*BHK)", txt, re.I)
    bhk   = bhk_m.group(1).strip() if bhk_m else ""

    price_m = re.search(
        r"₹\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore)(?:\s*[-–]\s*₹?\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore))?",
        txt, re.I
    )
    price = price_m.group(0).strip() if price_m else ""

    sqft_m = re.search(r"([\d,]+)\s*sq\.?\s*ft", txt, re.I)
    sqft   = sqft_m.group(1).replace(",", "") if sqft_m else ""

    amenity_keywords = [
        "Swimming Pool", "Pool", "Gym", "Lift", "Elevator", "Parking",
        "Garden", "Park", "Security", "CCTV", "Power Backup",
        "Clubhouse", "Club House", "Play Area", "Vastu", "Gated Community",
        "Metro", "Intercom", "Gas Pipeline", "Water Supply", "Fire Safety",
        "Children Play", "Jogging Track", "Multipurpose Hall",
        "Natural Light", "Grocery Store", "Utility", "Rainwater Harvesting",
    ]
    found     = [a for a in amenity_keywords if a.lower() in txt.lower()]
    amenities = ", ".join(dict.fromkeys(found))

    if not bhk and not price:
        return None
    return {"bhk": bhk, "sqft": sqft, "price": price, "amenities": amenities}


def parse_card(card):
    results = []
    try:
        units = card.query_selector_all("[data-testid*='property']")
        if units:
            for unit in units:
                prop = extract_fields(safe_text(unit))
                if prop:
                    results.append(prop)
        else:
            prop = extract_fields(safe_text(card))
            if prop:
                results.append(prop)
    except Exception:
        pass
    return results


# ── SCRAPE ALL PAGES FOR ONE LOCALITY ────────────────────────────────────────

def scrape_all_properties(page, apt_url, locality):
    all_props  = []
    page_num   = 1
    total      = None
    last_count = -1

    base_url = re.sub(r"[?&]page=\d+", "", apt_url)
    sep      = "&" if "?" in base_url else "?"

    while True:
        # ── ABSOLUTE HARD CAP — checked before every page load ───────────────
        if page_num > MAX_PAGES_ABSOLUTE:
            print(f"      🛑 Hit absolute page cap ({MAX_PAGES_ABSOLUTE}) — stopping")
            break

        current_url = base_url if page_num == 1 else f"{base_url}{sep}page={page_num}"
        print(f"      📄 Page {page_num}: {current_url[:80]}")

        if not load_page(page, current_url):
            break

        scroll_full(page)
        time.sleep(1)

        # Read total count on page 1 only
        if page_num == 1:
            for sel in ["div[class*='property']", "div[class*='result']",
                        "div[class*='count']", "div[class*='Count']"]:
                try:
                    count_el = page.query_selector(sel)
                    if count_el:
                        m = re.search(r"of\s*([\d,]+)", safe_text(count_el))
                        if m:
                            total = int(m.group(1).replace(",", ""))
                            print(f"      ℹ Total listings reported: {total}")
                            break
                except Exception:
                    pass

            # Sanity check: suspiciously large total vs actual cards on page 1
            if total and total > 500:
                cards_check = page.query_selector_all("article")
                if len(cards_check) < 30:
                    print(f"      ⚠ Reported {total} but only {len(cards_check)} "
                          f"cards on page 1 → overriding total to {len(cards_check)}")
                    total = len(cards_check)

        cards = page.query_selector_all("article")
        print(f"      ✔ {len(cards)} cards found")

        if not cards:
            print(f"      ⚠ No cards — stopping")
            break

        page_props = []
        for card in cards:
            page_props.extend(parse_card(card))

        print(f"      ✔ {len(page_props)} rows extracted  "
              f"(running total: {len(all_props) + len(page_props)})")

        if not page_props:
            break

        all_props.extend(page_props)

        if len(all_props) == last_count:
            break
        last_count = len(all_props)

        if total and len(all_props) >= total:
            print(f"      ✅ Collected all {total} listings")
            break

        # Dynamic cap from reported total, bounded by absolute cap
        if total:
            max_pages = min(math.ceil(total / 30) + 1, MAX_PAGES_ABSOLUTE)
            if page_num >= max_pages:
                print(f"      ✅ Reached max pages ({max_pages}) for total={total}")
                break

        page_num += 1
        time.sleep(1)

    print(f"      ✅ '{locality}' complete — {len(all_props)} apartments")
    return all_props


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb, ws = init_wb()
    done   = get_saved_localities(ws)
    if done:
        print(f"↩  Skipping {len(done)} already-saved localities\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=30)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
        )
        page = ctx.new_page()

        for pg in range(1, TOTAL_PAGES + 1):
            table_url = f"{BASE_URL}?page={pg}"
            print(f"\n{'='*65}")
            print(f"  TABLE PAGE {pg}/{TOTAL_PAGES}")
            print(f"{'='*65}")

            if not load_page(page, table_url):
                continue

            scroll_full(page)
            time.sleep(1)

            locality_data = get_locality_rows(page)
            if not locality_data:
                print(f"  ⚠ No localities on page {pg}")
                continue

            for ld in locality_data:
                print(f"   • {ld['locality']:25} {ld['avg_price']:12}  "
                      f"{ld['p_min']} – {ld['p_max']}")

            for ld in locality_data:
                locality = ld["locality"]
                if locality in done:
                    print(f"\n  ⏭  Skip: {locality}")
                    continue

                print(f"\n  📍 '{locality}'")

                props = []
                if ld["prop_url"]:
                    apt_url = get_apartment_url(page, ld["prop_url"])
                    if apt_url:
                        props = scrape_all_properties(page, apt_url, locality)
                    else:
                        print(f"   ⚠ Could not get Apartment URL")

                    load_page(page, table_url)
                    scroll_full(page)
                else:
                    print(f"   ⚠ No listing URL")

                save_locality(wb, ws, locality,
                              ld["avg_price"], ld["p_min"], ld["p_max"], props)
                done.add(locality)
                time.sleep(1.5)

        browser.close()

    print(f"\n{'='*65}")
    print(f"✅  ALL DONE!  Total rows: {ws.max_row - 1}  →  {OUTPUT_FILE}")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()