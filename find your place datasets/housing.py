"""
Housing.com - Hyderabad APARTMENT Scraper (Final)
==================================================
Exact selectors confirmed from inspect_listings.py:
  - Property cards  : article  (60 per page)
  - Individual unit : [data-testid*='property']  (30 per page)
  - Total count div : div[class*='property']  e.g. "Showing 31-60 of 1072"

Saves apartment.xlsx after EVERY locality. Auto-resumes on re-run.

Setup:
    pip install playwright openpyxl
    playwright install chromium

Run:
    python apartment_scraper.py
"""

import re
import time
import os
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ───────────────────────────────────────────────────────────────────
BASE_URL    = "https://housing.com/price-trends/property-rates-for-buy-in-hyderabad_telangana-P679xe73u28050522"
TOTAL_PAGES = 17
OUTPUT_FILE = "apartment.xlsx"
HEADLESS    = False

# ── EXCEL ─────────────────────────────────────────────────────────────────────
COLUMNS    = ["Locality", "Avg Price/Sqft", "Price Min", "Price Max",
              "BHK", "Sqft", "Price", "Amenities"]
COL_WIDTHS = [22, 18, 16, 16, 10, 12, 15, 55]
HDR_FILL   = PatternFill("solid", fgColor="4B0082")
HDR_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ALT_FILL   = PatternFill("solid", fgColor="F3EEFF")
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def init_wb():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        print(f"↩  Resuming {OUTPUT_FILE}  ({ws.max_row - 1} rows saved)")
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apartment"
    for ci, (h, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"✔  Created {OUTPUT_FILE}")
    return wb, ws


def save_locality(wb, ws, locality, avg_price, p_min, p_max, props):
    if not props:
        props = [{"bhk": "", "sqft": "", "price": "", "amenities": ""}]
    for prop in props:
        r = ws.max_row + 1
        vals = [locality, avg_price, p_min, p_max,
                prop["bhk"], prop["sqft"], prop["price"], prop["amenities"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.border = BORDER
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                c.fill = ALT_FILL
    wb.save(OUTPUT_FILE)
    print(f"   💾 Saved {len(props)} rows for '{locality}'  (total: {ws.max_row - 1})")


def get_saved_localities(ws):
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    return done


def safe_text(el):
    try:
        return el.inner_text().strip() if el else ""
    except:
        return ""


def scroll_full(page):
    for pct in [0.3, 0.6, 1.0]:
        page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pct})")
        time.sleep(0.8)


def load_page(page, url, retries=3):
    for attempt in range(retries):
        try:
            page.goto(url, timeout=50000, wait_until="domcontentloaded")
            time.sleep(4)
            return True
        except PWTimeout:
            print(f"   Timeout attempt {attempt+1}/{retries}, retrying...")
            time.sleep(3)
    print(f"   ✗ Failed: {url}")
    return False


# ── TABLE: get locality rows ──────────────────────────────────────────────────

def get_locality_rows(page):
    rows_data = []

    # Exact class from inspector output
    rows = page.query_selector_all("div.css-1s17y02")

    # Fallback: any div containing a price-trend locality link
    if not rows:
        rows = [r for r in page.query_selector_all("div[class^='css-']")
                if r.query_selector("a[href*='property-rates']")]

    print(f"   Found {len(rows)} locality rows")

    for row in rows:
        try:
            name_el = (row.query_selector("a.css-673lf3") or
                       row.query_selector("a[href*='property-rates']") or
                       row.query_selector("a"))
            locality = safe_text(name_el)
            if not locality:
                continue

            price_el  = (row.query_selector("span.css-69n8oe") or
                         row.query_selector("span:nth-child(2)"))
            range_el  = (row.query_selector("span.css-5sq9yq") or
                         row.query_selector("span:nth-child(3)"))

            avg_price   = safe_text(price_el)
            price_range = safe_text(range_el)
            parts  = re.split(r"\s*[-–]\s*", price_range)
            p_min  = parts[0].strip() if parts else ""
            p_max  = parts[1].strip() if len(parts) > 1 else ""

            # "See X Properties" link → /in/buy/hyderabad/miyapur
            link_el = (row.query_selector("span.css-15j6032 a") or
                       row.query_selector("a[href*='/in/buy/']"))
            href = link_el.get_attribute("href") if link_el else ""
            prop_url = ("https://housing.com" + href if href.startswith("/")
                        else href if href.startswith("http") else "")

            rows_data.append({
                "locality":  locality,
                "avg_price": avg_price,
                "p_min":     p_min,
                "p_max":     p_max,
                "prop_url":  prop_url,
            })
        except Exception as e:
            print(f"   Row parse error: {e}")
    return rows_data


# ── LISTINGS: parse one article card ─────────────────────────────────────────

def parse_card(card):
    """
    Extract BHK, sqft, price, amenities from one article card.
    The article contains multiple [data-testid*='property'] sub-units
    (individual flats listed under a project). Parse each sub-unit separately.
    """
    results = []

    # Try sub-units first (individual listings inside a project card)
    units = card.query_selector_all("[data-testid*='property']")

    if units:
        for unit in units:
            txt = safe_text(unit)
            if not txt:
                continue
            prop = extract_fields(txt)
            if prop:
                results.append(prop)
    else:
        # Treat whole card as one listing
        txt = safe_text(card)
        prop = extract_fields(txt)
        if prop:
            results.append(prop)

    return results


def extract_fields(txt):
    """Parse BHK, sqft, price, amenities from raw text block."""
    if not txt or len(txt) < 15:
        return None

    # BHK — e.g. "2 BHK", "2.5 BHK", "1, 2 BHK"
    bhk_m = re.search(r"([\d.,\s&]+\s*BHK)", txt, re.I)
    bhk   = bhk_m.group(1).strip() if bhk_m else ""

    # Price — ₹XX L / ₹XX Cr / ₹XX.XX L-XX.XX L
    price_m = re.search(
        r"₹\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore)(?:\s*[-–]\s*₹?\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore))?",
        txt, re.I
    )
    price = price_m.group(0).strip() if price_m else ""

    # Sqft — "1180 sq.ft" / "1,180 sqft"
    sqft_m = re.search(r"([\d,]+)\s*sq\.?\s*ft", txt, re.I)
    sqft   = sqft_m.group(1).replace(",", "") if sqft_m else ""

    # Amenities
    amenity_keywords = [
        "Swimming Pool", "Pool", "Gym", "Lift", "Elevator", "Parking",
        "Garden", "Park", "Security", "CCTV", "Power Backup",
        "Clubhouse", "Club House", "Play Area", "Vastu", "Gated Community",
        "Metro", "Intercom", "Gas Pipeline", "Water Supply", "Fire Safety",
        "Children Play", "Jogging Track", "Multipurpose Hall",
    ]
    found = [a for a in amenity_keywords if a.lower() in txt.lower()]
    amenities = ", ".join(dict.fromkeys(found))

    if not bhk and not price:
        return None
    return {"bhk": bhk, "sqft": sqft, "price": price, "amenities": amenities}


# ── LISTINGS: scrape all pages for one locality ───────────────────────────────

def scrape_all_properties(page, prop_url, locality):
    all_props = []
    current_url = prop_url
    listing_page = 1
    visited = set()

    while current_url and current_url not in visited and listing_page <= 200:
        visited.add(current_url)
        print(f"      📄 Page {listing_page}: {current_url[:80]}")

        if not load_page(page, current_url):
            break

        scroll_full(page)
        time.sleep(1)

        # ── How many total properties? ────────────────────────────────
        total = 0
        try:
            count_el = page.query_selector("div[class*='property']")
            if count_el:
                m = re.search(r"of\s*([\d,]+)", safe_text(count_el))
                if m:
                    total = int(m.group(1).replace(",", ""))
                    if listing_page == 1:
                        print(f"      ℹ Total listings for '{locality}': {total}")
        except Exception:
            pass

        # ── Grab article cards ────────────────────────────────────────
        cards = page.query_selector_all("article")
        print(f"      ✔ {len(cards)} article cards found")

        if not cards:
            print(f"      ⚠ No cards — stopping")
            break

        page_count = 0
        for card in cards:
            try:
                props = parse_card(card)
                all_props.extend(props)
                page_count += len(props)
            except Exception:
                continue

        print(f"      ✔ Extracted {page_count} rows  (running total: {len(all_props)})")

        # ── Next page ─────────────────────────────────────────────────
        next_url = None
        try:
            # rel=next link
            nxt = page.query_selector("a[rel='next']")
            if nxt:
                href = nxt.get_attribute("href") or ""
                next_url = ("https://housing.com" + href
                            if href.startswith("/") else href)

            # Active page number
            if not next_url:
                active = (page.query_selector("[aria-current='page']") or
                          page.query_selector("[class*='activePage']") or
                          page.query_selector("[class*='active'][class*='page']"))
                if active:
                    num = safe_text(active)
                    if num.isdigit():
                        base = re.sub(r"[?&]page=\d+", "", current_url)
                        sep = "&" if "?" in base else "?"
                        next_url = f"{base}{sep}page={int(num)+1}"

            # Always try page=2 on first listing page
            if not next_url and listing_page == 1:
                base = re.sub(r"[?&]page=\d+", "", current_url)
                sep = "&" if "?" in base else "?"
                next_url = f"{base}{sep}page=2"

        except Exception:
            pass

        # Stop if we've got all properties
        if total and len(all_props) >= total:
            print(f"      ✅ Collected all {total} properties")
            break

        if next_url and next_url not in visited:
            current_url = next_url
            listing_page += 1
        else:
            break

    print(f"      ✅ '{locality}' done — {len(all_props)} properties total")
    return all_props


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb, ws = init_wb()
    done = get_saved_localities(ws)
    if done:
        print(f"↩  Skipping {len(done)} already-saved localities\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=30)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
        )
        page = ctx.new_page()

        for pg in range(1, TOTAL_PAGES + 1):
            table_url = f"{BASE_URL}?page={pg}"
            print(f"\n{'='*65}")
            print(f"  TABLE PAGE {pg}/{TOTAL_PAGES}")
            print(f"{'='*65}")

            if not load_page(page, table_url):
                continue

            # Click Apartment tab
            try:
                for tab in page.query_selector_all(
                    "button[role='tab'], div[role='tab'], [class*='tab']"
                ):
                    if safe_text(tab).strip().lower() in ("apartment", "apartments"):
                        tab.click()
                        print(f"  ✔ Clicked 'Apartment' tab")
                        time.sleep(2)
                        break
            except Exception:
                pass

            scroll_full(page)
            time.sleep(1)

            locality_data = get_locality_rows(page)
            if not locality_data:
                print(f"  ⚠ No localities found on table page {pg}")
                continue

            for ld in locality_data:
                print(f"   • {ld['locality']:25} {ld['avg_price']:12}  "
                      f"{ld['p_min']} – {ld['p_max']}  🔗 ...{ld['prop_url'][-40:]}")

            for ld in locality_data:
                locality = ld["locality"]
                if locality in done:
                    print(f"\n  ⏭  Skip: {locality}")
                    continue

                print(f"\n  📍 '{locality}'")

                props = []
                if ld["prop_url"]:
                    props = scrape_all_properties(page, ld["prop_url"], locality)
                    load_page(page, table_url)
                    scroll_full(page)
                else:
                    print(f"   ⚠ No listing URL — saving locality info only")

                save_locality(wb, ws, locality,
                              ld["avg_price"], ld["p_min"], ld["p_max"], props)
                done.add(locality)
                time.sleep(1.5)

        browser.close()

    print(f"\n{'='*65}")
    print(f"✅  ALL DONE!  Total rows: {ws.max_row - 1}  →  {OUTPUT_FILE}")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()