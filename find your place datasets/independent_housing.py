"""
Housing.com - Hyderabad INDEPENDENT HOUSE Scraper
===================================================
Same approach as villa_scraper.py:
  1. Opens each locality listing page e.g. /in/buy/hyderabad/bandlaguda
  2. Clicks "Property Type" dropdown (div.css-r9ub9f)
  3. Finds the "Independent House" <a> link inside the dropdown
     e.g. <a href="/in/buy/hyderabad/house-bandlaguda">Independent House</a>
  4. Navigates to that URL and scrapes ALL results with pagination
  5. Saves to independent_house.xlsx after every locality

Setup:
    pip install playwright openpyxl
    playwright install chromium

Run:
    python independent_house_scraper.py
"""

import re
import time
import os
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ───────────────────────────────────────────────────────────────────
BASE_URL    = "https://housing.com/price-trends/property-rates-for-buy-in-hyderabad_telangana-P679xe73u28050522"
TOTAL_PAGES = 17
OUTPUT_FILE = "independent_house.xlsx"
HEADLESS    = False

# ── EXCEL ─────────────────────────────────────────────────────────────────────
COLUMNS    = ["Locality", "Avg Price/Sqft", "Price Min", "Price Max",
              "BHK", "Sqft", "Price", "Amenities"]
COL_WIDTHS = [25, 18, 16, 16, 10, 12, 15, 55]
HDR_FILL   = PatternFill("solid", fgColor="1A237E")   # deep blue for Ind. House
HDR_FONT   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ALT_FILL   = PatternFill("solid", fgColor="E8EAF6")
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def init_wb():
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        print(f"↩  Resuming {OUTPUT_FILE}  ({ws.max_row - 1} rows saved)")
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Independent House"
    for ci, (h, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"✔  Created {OUTPUT_FILE}")
    return wb, ws


def save_locality(wb, ws, locality, avg_price, p_min, p_max, props):
    if not props:
        props = [{"bhk": "", "sqft": "", "price": "", "amenities": ""}]
    for prop in props:
        r = ws.max_row + 1
        vals = [locality, avg_price, p_min, p_max,
                prop["bhk"], prop["sqft"], prop["price"], prop["amenities"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.border = BORDER
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                c.fill = ALT_FILL
    wb.save(OUTPUT_FILE)
    print(f"   💾 Saved {len(props)} rows for '{locality}'  (total: {ws.max_row - 1})")


def get_saved_localities(ws):
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            done.add(str(row[0]).strip())
    return done


def safe_text(el):
    try:
        return el.inner_text().strip() if el else ""
    except:
        return ""


def scroll_full(page):
    for pct in [0.3, 0.6, 1.0]:
        page.evaluate(f"window.scrollTo(0, document.body.scrollHeight * {pct})")
        time.sleep(0.8)


def load_page(page, url, retries=3):
    for attempt in range(retries):
        try:
            page.goto(url, timeout=50000, wait_until="domcontentloaded")
            time.sleep(4)
            return True
        except PWTimeout:
            print(f"   Timeout attempt {attempt+1}/{retries}, retrying...")
            time.sleep(3)
    print(f"   ✗ Failed: {url}")
    return False


# ── GET INDEPENDENT HOUSE URL FOR A LOCALITY ─────────────────────────────────

def get_independent_house_url(page, locality_url):
    """
    Open the locality listing page, click Property Type dropdown,
    find the 'Independent House' <a> link, return its full URL.
    e.g. /in/buy/hyderabad/house-bandlaguda
      → https://housing.com/in/buy/hyderabad/house-bandlaguda

    From inspect output:
      - Dropdown button:  div.css-r9ub9f  text='Property Type'
      - Option wrapper:   div.option.css-18o4tmu  text='Independent House'
      - Link inside:      <a href="/in/buy/hyderabad/house-...">Independent House</a>
    """
    print(f"      🔍 Finding Independent House URL from: {locality_url[:70]}")

    if not load_page(page, locality_url):
        return None

    # Click "Property Type" dropdown — confirmed class: css-r9ub9f
    clicked = page.evaluate("""
        () => {
            const els = document.querySelectorAll('div.css-r9ub9f');
            for (const el of els) {
                if (el.innerText.trim() === 'Property Type') {
                    el.click();
                    return true;
                }
            }
            // Fallback: text walker
            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
            let node;
            while (node = walker.nextNode()) {
                if (node.textContent.trim() === 'Property Type') {
                    node.parentElement.click();
                    return true;
                }
            }
            return false;
        }
    """)

    if clicked:
        print(f"      ✔ Opened Property Type dropdown")
    else:
        print(f"      ⚠ Could not open Property Type dropdown")
        return None

    time.sleep(1.5)

    # Find the "Independent House" <a> href inside dropdown
    # From inspect: div.option.css-18o4tmu > span.css-0 > a
    ind_house_href = page.evaluate("""
        () => {
            // Primary: look inside div.option elements (confirmed from inspect)
            const options = document.querySelectorAll('div.option, div[class*="option"]');
            for (const opt of options) {
                if (opt.innerText.trim() === 'Independent House') {
                    const link = opt.querySelector('a');
                    if (link) return link.getAttribute('href');
                }
            }
            // Fallback: find <a> whose exact text is "Independent House"
            const links = document.querySelectorAll('a');
            for (const a of links) {
                if (a.innerText.trim() === 'Independent House') {
                    return a.getAttribute('href');
                }
            }
            return null;
        }
    """)

    if ind_house_href:
        full_url = ("https://housing.com" + ind_house_href
                    if ind_house_href.startswith("/") else ind_house_href)
        print(f"      ✔ Independent House URL: {full_url}")
        return full_url
    else:
        print(f"      ⚠ 'Independent House' link not found in dropdown")
        page.keyboard.press("Escape")
        return None


# ── TABLE: get locality rows ──────────────────────────────────────────────────

def get_locality_rows(page):
    rows_data = []
    rows = page.query_selector_all("div.css-1s17y02")
    if not rows:
        rows = [r for r in page.query_selector_all("div[class^='css-']")
                if r.query_selector("a[href*='property-rates']")]
    print(f"   Found {len(rows)} locality rows")

    for row in rows:
        try:
            name_el  = (row.query_selector("a.css-673lf3") or
                        row.query_selector("a[href*='property-rates']") or
                        row.query_selector("a"))
            locality = safe_text(name_el)
            if not locality:
                continue

            price_el    = (row.query_selector("span.css-69n8oe") or
                           row.query_selector("span:nth-child(2)"))
            range_el    = (row.query_selector("span.css-5sq9yq") or
                           row.query_selector("span:nth-child(3)"))
            avg_price   = safe_text(price_el)
            price_range = safe_text(range_el)
            parts = re.split(r"\s*[-–]\s*", price_range)
            p_min = parts[0].strip() if parts else ""
            p_max = parts[1].strip() if len(parts) > 1 else ""

            link_el  = (row.query_selector("span.css-15j6032 a") or
                        row.query_selector("a[href*='/in/buy/']"))
            href     = link_el.get_attribute("href") if link_el else ""
            prop_url = ("https://housing.com" + href if href.startswith("/")
                        else href if href.startswith("http") else "")

            rows_data.append({
                "locality":  locality,
                "avg_price": avg_price,
                "p_min":     p_min,
                "p_max":     p_max,
                "prop_url":  prop_url,
            })
        except Exception as e:
            print(f"   Row parse error: {e}")
    return rows_data


# ── PARSE CARDS ───────────────────────────────────────────────────────────────

def extract_fields(txt):
    if not txt or len(txt) < 15:
        return None

    bhk_m = re.search(r"([\d.,\s&]+\s*BHK)", txt, re.I)
    bhk   = bhk_m.group(1).strip() if bhk_m else ""

    price_m = re.search(
        r"₹\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore)(?:\s*[-–]\s*₹?\s*[\d,.]+\s*(?:L|Lakh|Cr|Crore))?",
        txt, re.I
    )
    price = price_m.group(0).strip() if price_m else ""

    sqft_m = re.search(r"([\d,]+)\s*sq\.?\s*ft", txt, re.I)
    sqft   = sqft_m.group(1).replace(",", "") if sqft_m else ""

    amenity_keywords = [
        "Swimming Pool", "Pool", "Gym", "Lift", "Elevator", "Parking",
        "Garden", "Park", "Security", "CCTV", "Power Backup",
        "Clubhouse", "Club House", "Play Area", "Vastu", "Gated Community",
        "Metro", "Intercom", "Gas Pipeline", "Water Supply", "Fire Safety",
        "Children Play", "Jogging Track", "Natural Light", "Open Space",
        "Rainwater Harvesting", "Terrace", "Duplex", "Sports Facility",
        "Private Pool", "Private Garden", "Home Theatre",
        "Modular Kitchen", "Solar Panel",
    ]
    found     = [a for a in amenity_keywords if a.lower() in txt.lower()]
    amenities = ", ".join(dict.fromkeys(found))

    if not bhk and not price:
        return None
    return {"bhk": bhk, "sqft": sqft, "price": price, "amenities": amenities}


def parse_card(card):
    results = []
    try:
        units = card.query_selector_all("[data-testid*='property']")
        if units:
            for unit in units:
                prop = extract_fields(safe_text(unit))
                if prop:
                    results.append(prop)
        else:
            prop = extract_fields(safe_text(card))
            if prop:
                results.append(prop)
    except Exception:
        pass
    return results


# ── SCRAPE ALL PROPERTIES FOR ONE LOCALITY ────────────────────────────────────

def scrape_all_properties(page, ind_house_url, locality):
    all_props  = []
    page_num   = 1
    total      = None
    last_count = -1

    base_url = re.sub(r"[?&]page=\d+", "", ind_house_url)
    sep      = "&" if "?" in base_url else "?"

    while True:
        current_url = base_url if page_num == 1 else f"{base_url}{sep}page={page_num}"
        print(f"      📄 Page {page_num}: {current_url[:80]}")

        if not load_page(page, current_url):
            break

        scroll_full(page)
        time.sleep(1)

        if total is None:
            try:
                for sel in ["div[class*='property']", "div[class*='result']",
                            "div[class*='count']", "div[class*='Count']"]:
                    count_el = page.query_selector(sel)
                    if count_el:
                        m = re.search(r"of\s*([\d,]+)", safe_text(count_el))
                        if m:
                            total = int(m.group(1).replace(",", ""))
                            print(f"      ℹ Total Independent House listings: {total}")
                            break
            except Exception:
                pass

        cards = page.query_selector_all("article")
        print(f"      ✔ {len(cards)} cards found")

        if not cards:
            print(f"      ⚠ No cards — stopping")
            break

        page_props = []
        for card in cards:
            page_props.extend(parse_card(card))

        print(f"      ✔ {len(page_props)} rows extracted  "
              f"(running total: {len(all_props) + len(page_props)})")

        if not page_props:
            break

        all_props.extend(page_props)

        if len(all_props) == last_count:
            break
        last_count = len(all_props)

        if total and len(all_props) >= total:
            print(f"      ✅ Collected all {total} properties")
            break

        if total:
            max_pages = (total // 30) + 2
            if page_num >= max_pages:
                break

        page_num += 1
        time.sleep(1)

    print(f"      ✅ '{locality}' complete — {len(all_props)} Independent House properties")
    return all_props


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb, ws = init_wb()
    done   = get_saved_localities(ws)
    if done:
        print(f"↩  Skipping {len(done)} already-saved localities\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=30)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1400, "height": 900},
        )
        page = ctx.new_page()

        for pg in range(1, TOTAL_PAGES + 1):
            table_url = f"{BASE_URL}?page={pg}"
            print(f"\n{'='*65}")
            print(f"  TABLE PAGE {pg}/{TOTAL_PAGES}")
            print(f"{'='*65}")

            if not load_page(page, table_url):
                continue

            # Click "Independent House" tab on price-trends page (if tab exists)
            page.evaluate("""
                () => {
                    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
                    let node;
                    while (node = walker.nextNode()) {
                        if (node.textContent.trim() === 'Independent House') {
                            node.parentElement.click();
                            return true;
                        }
                    }
                    return false;
                }
            """)
            time.sleep(2)
            scroll_full(page)

            locality_data = get_locality_rows(page)
            if not locality_data:
                print(f"  ⚠ No localities found on page {pg}")
                continue

            for ld in locality_data:
                print(f"   • {ld['locality']:25} {ld['avg_price']:12}  "
                      f"{ld['p_min']} – {ld['p_max']}")

            for ld in locality_data:
                locality = ld["locality"]
                if locality in done:
                    print(f"\n  ⏭  Skip: {locality}")
                    continue

                print(f"\n  📍 '{locality}'")

                props = []
                if ld["prop_url"]:
                    # Get Independent House-specific URL from dropdown
                    ind_house_url = get_independent_house_url(page, ld["prop_url"])

                    if ind_house_url:
                        props = scrape_all_properties(page, ind_house_url, locality)
                    else:
                        print(f"   ⚠ Could not get Independent House URL — skipping")

                    # Return to table page and re-click Independent House tab
                    load_page(page, table_url)
                    page.evaluate("""
                        () => {
                            const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
                            let node;
                            while (node = walker.nextNode()) {
                                if (node.textContent.trim() === 'Independent House') {
                                    node.parentElement.click();
                                    return true;
                                }
                            }
                        }
                    """)
                    time.sleep(2)
                    scroll_full(page)
                else:
                    print(f"   ⚠ No listing URL")

                save_locality(wb, ws, locality,
                              ld["avg_price"], ld["p_min"], ld["p_max"], props)
                done.add(locality)
                time.sleep(1.5)

        browser.close()

    print(f"\n{'='*65}")
    print(f"✅  ALL DONE!  Total rows: {ws.max_row - 1}  →  {OUTPUT_FILE}")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()